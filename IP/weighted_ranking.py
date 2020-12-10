import xlsxwriter
import openpyxl as px

class Weight_association:
    def __init__(self, spice, disease , count ):
        self.spice = spice
        self.disease = disease
        self.count = count
        # self.ws_score = ws_score

class Weight_association_ranking:
    def __init__(self, spice, disease , sum_count):
        self.spice = spice
        self.disease = disease
        self.sum_count = sum_count
        # self.sum_ws_score = sum_ws_score


workbook = xlsxwriter.Workbook('Maximal_Bicliques_weighted.xlsx')
worksheet = workbook.add_worksheet()
row = 0
column = 0
row1 = 0
bicliques_file = open('Maximal_Biclique_final_mod.txt', 'r')
bi_lines = bicliques_file.readlines()

my_file = open('level-3-associations.tsv', 'r')
# my_weighted_file = open('most-associations-each-level.tsv', 'r')
lines = my_file.readlines()
# my_lines = my_weighted_file.readlines()
lines.pop(0)
# my_lines.pop(0)

wa_ob_list = []
for line in lines:
    # for my_line in my_lines:
        temp_list = line.split('\t')
        # temp_list1 = my_line.split('\t')
        # if temp_list[2].strip() == temp_list1[2].strip() and temp_list[4].strip() == temp_list1[4].strip():
        wa = Weight_association(temp_list[1].strip(), temp_list[2].strip(), temp_list[3].strip())
        wa_ob_list.append(wa)

temp_list3 = []
for bi_line in bi_lines:
    if bi_line.strip():
        temp_list3 = bi_line.split(':')
        temp_list4 = []
        temp_list5 = []
        # print(temp_list[0])
        temp_list4 = temp_list3[0].split(';')
        temp_list14 = list(set(temp_list4))
        # print(temp_list[1])
        temp_list5 = temp_list3[1].split(';')
        temp_list5 = list(set(temp_list5))
        sum_count = 0
        # ws_spectrum_sum = 0
        row = row1
        start_idx = row
        for spices in temp_list4:
            worksheet.write(row, column, spices)
            row+=1
            for diseases in temp_list5:
                for wao in wa_ob_list:
                    if wao.spice == spices and wao.disease == diseases:
                        count = int(wao.count)
                        sum_count = sum_count + count
                        # ws_spectrum = float(wao.ws_score)
                        # ws_spectrum_sum = ws_spectrum_sum + ws_spectrum
                        worksheet.write(row1, column+1, diseases)
                        row1 += 1
        end_idx = row1 + 2
        for i in range(start_idx, end_idx):
            worksheet.write(i, column + 2, sum_count)
            # worksheet.write(i, column + 3, ws_spectrum_sum)
        row1 += 2
workbook.close()
#
# W = px.load_workbook('Maximal_Bicliques_weighted.xlsx', use_iterators = True)
# p = W.get_sheet_by_name(name = 'Sheet1')
#
# a=[]
# spice = []
# disease = []
# for row in p.iter_rows():
#     if(row) and (row[1]):
#         spice.append(row[0])
#         disease.append(row[1])
#         wr = Weight_association_ranking(row[0], row[1])


